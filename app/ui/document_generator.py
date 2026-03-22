import os
import re
import sys 
import shutil
import math
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side
from ..db import get_conn, query_all, get_next_db_serial, update_db_serial
from .utils import resource_path
from openpyxl.drawing.image import Image as XLImage

# 제품명 약어 매핑
ABBR_MAP = {
    "B10000805055": "MOD 29",
    "B10000850333": "38-101A",
    "B10000850460": "32-460",
    "B10000850490": "32-450",
    "B10000851180": "CRPS",
    "B10000852308": "MOD 781",
    "B10000852323": "MOD 781B"
}


def get_output_dir(sub_folder: str):
    """지정된 서브 폴더의 출력 경로를 반환합니다."""
    # 실행 파일(또는 스크립트) 위치 기준 'Output_doc/{sub_folder}'
    if getattr(sys, 'frozen', False):
        base_path = os.path.dirname(sys.executable)
    else:
        # run.py가 있는 폴더 찾기
        base_path = os.path.dirname(os.path.abspath(__file__))
        base_path = str(Path(base_path).parent.parent)

    target_dir = Path(base_path) / "Output_doc" / sub_folder
    target_dir.mkdir(parents=True, exist_ok=True)
    return target_dir

def get_oa_output_dir():
    return get_output_dir("Order Acknowledgement")

def get_po_output_dir():
    return get_output_dir("Purchase Order")

def insert_transparent_stamp(ws, fixed_anchor=None):
    """
    워크시트의 헤더 영역(1~15행)에 있는 기존 이미지를 찾아 
    투명 배경이 적용된 'app/ui/stamp.png'로 교체합니다.
    (기존 이미지가 없으면 J4에 삽입 시도)
    """
    try:
        stamp_path = resource_path("app/ui/stamp.png")
        if not os.path.exists(stamp_path):
            print(f"Stamp image not found: {stamp_path}")
            return

        # 1. 위치 결정
        target_anchor = "J4" # 기본값

        if fixed_anchor:
            target_anchor = fixed_anchor
        else:
            # 기존 헤더 이미지 찾기 로직
            images_to_remove = []
            
            # ws._images는 리스트이므로 순회
            if hasattr(ws, "_images"):
                for img in ws._images:
                    # 앵커 위치 확인 (_from.row is 0-indexed)
                    try:
                        r = img.anchor._from.row
                        c = img.anchor._from.col
                        # 헤더 영역(대략 15행 이내)에 있는 이미지라고 가정
                        if r < 15:
                            # 앵커를 문자로 변환 (예: 4,9 -> J5)
                            # openpyxl col is 0-indexed here? usually.
                            # utils.get_column_letter is 1-indexed.
                            from openpyxl.utils import get_column_letter
                            col_letter = get_column_letter(c + 1)
                            target_anchor = f"{col_letter}{r + 1}"
                            images_to_remove.append(img)
                            # 첫 번째 발견된 헤더 이미지만 교체 (보통 도장 하나임)
                            break 
                    except:
                        continue
            
            # 2. 기존 이미지 제거
            for img in images_to_remove:
                ws._images.remove(img)
            
        # 3. 새 이미지 삽입
        new_img = XLImage(stamp_path)
        # 사이즈 조정 (1.98cm x 1.93cm) -> 75px x 73px
        new_img.width = 75
        new_img.height = 73
        
        # 위치 조정 (좌측 35px, 아래 20px)
        # 전략: 'J4' 대신 'I4'에 이미지 삽입 후, 오프셋으로 우측 이동하여 위치 미세 조정
        # 'I' 열의 너비가 보통 64px(8.38) ~ 70px 정도.
        # J4에서 좌측 35px -> I4의 우측 끝에서 35px 안쪽?
        # 대략 I4 셀의 시작점에서 30px 정도 띄우면 J4 기준 좌측 35px 부근에 위치한다고 가정.
        # (정확한 컬럼 너비를 모르므로 근사치 값 적용)
        
        from openpyxl.utils import get_column_letter, column_index_from_string
        from openpyxl.utils.units import pixels_to_EMU
        import re
        
        match = re.match(r"([A-Z]+)(\d+)", target_anchor)
        if match:
            col_str, row_str = match.groups()
            col_idx = column_index_from_string(col_str) # 1-based (J=10)
            row_idx = int(row_str)
            
            # 1칸 왼쪽 컬럼 (I=9)
            new_col_idx = max(1, col_idx - 1)
            new_col_str = get_column_letter(new_col_idx)
            new_anchor = f"{new_col_str}{row_idx}"
            
            # 이미지 먼저 추가 (Anchor 객체 자동 생성)
            ws.add_image(new_img, new_anchor)
            
            # 오프셋 상세 조정
            if new_img.anchor:
                try:
                    marker = new_img.anchor._from
                    
                    # [위치 수정 가이드]
                    # colOff: 가로, rowOff: 세로 (1px = 9525 EMU)
                    
                    if fixed_anchor == "E10":
                        # 발주서(PO) 전용: E10 (권성관 대표 오른쪽)
                        marker.colOff = pixels_to_EMU(10) 
                        marker.rowOff = pixels_to_EMU(5) 
                        
                    # 기존 로직 (상대적 위치)
                    elif not fixed_anchor and target_anchor.startswith("I"):
                         marker.colOff = pixels_to_EMU(240)
                         marker.rowOff = pixels_to_EMU(120)
                         
                except:
                    pass
        elif not fixed_anchor:
             # Fallback
             ws.add_image(new_img, target_anchor)
        else:
             ws.add_image(new_img, target_anchor)
        
    except Exception as e:
        print(f"Stamp Insertion Error: {e}")

def get_next_oa_serial(year_str: str, order_id: int = None) -> str:
    """OA 일련번호 (DB 기반 또는 주문 순번)"""
    if order_id:
        try:
            conn = get_conn()
            cur = conn.cursor()
            
            # 해당 주문의 날짜 조회
            cur.execute("SELECT order_dt FROM orders WHERE id = ?", (order_id,))
            row = cur.fetchone()
            if row and row[0]:
                target_dt = row[0] # YYYY-MM-DD
                
                # 해당 연도 내에서 이 주문의 순번 계산 (날짜순 -> ID순)
                # target_dt와 같은 연도인 주문들 중
                # 1. 날짜가 더 빠른 주문
                # 2. 날짜가 같으면 ID가 작거나 같은 주문
                # 의 개수 = 순위
                sql = """
                    SELECT COUNT(*) 
                    FROM orders
                    WHERE strftime('%Y', order_dt) = ?
                      AND (order_dt < ? OR (order_dt = ? AND id <= ?))
                """
                cur.execute(sql, (year_str, target_dt, target_dt, order_id))
                rank = cur.fetchone()[0]
                conn.close()
                return f"{rank:03d}"
            conn.close()
        except Exception as e:
            print(f"OA Serial Rank Calc Error: {e}")

    # Fallback: 기존 로직 (다음 번호)
    next_val = get_next_db_serial("OA", year_str)
    return f"{next_val:03d}"

def get_delivery_serial_rank(delivery_id: int) -> str:
    """납품 일련번호 (연도별 누적 주문 연결 건수 기반 순번)"""
    try:
        conn = get_conn()
        cur = conn.cursor()
        
        # 납품일 조회
        cur.execute("SELECT ship_datetime, invoice_no FROM deliveries WHERE id = ?", (delivery_id,))
        row = cur.fetchone()
        
        if row and row[0]:
            ship_dt = row[0] # YYYY-MM-DD or YYYY-MM-DD HH:MM:SS
            # 연도 추출 (YYYY)
            target_year = ship_dt[:4]
            
            # 해당 연도 내, 현재 납품보다 이전 시점(또는 같은 시점 이전 ID)의 납품들에 연결된
            # [유니크한 주문 개수]를 카운트
            # 즉, 이전에 생성됐어야 할 인보이스의 총 개수
            sql = """
                SELECT COUNT(*)
                FROM (
                    SELECT DISTINCT di.delivery_id, di.order_id
                    FROM delivery_items di
                    JOIN deliveries d ON di.delivery_id = d.id
                    WHERE strftime('%Y', d.ship_datetime) = ?
                      AND di.order_id IS NOT NULL
                      AND (d.ship_datetime < ? OR (d.ship_datetime = ? AND d.id < ?))
                )
            """
            cur.execute(sql, (target_year, ship_dt, ship_dt, delivery_id))
            prev_invoice_count = cur.fetchone()[0]
            
            # 현재 납품의 시작 번호 = 이전 개수 + 1
            rank = prev_invoice_count + 1
            
            conn.close()
            return f"{rank:03d}"
            
        conn.close()
        return "001"
    except Exception as e:
        print(f"Delivery Serial Rank Error: {e}")
        return "001"

def get_next_po_serial(year_str: str, prefix_type: str = "MF") -> str:
    """
    발주서 파일(Purchase Order_...)을 스캔하여 다음 일련번호 반환.
    prefix_type: "MF" 또는 "TO"
    MF 패턴: MF{YYMMDD}-{XXX}
    TO 패턴: TO{YYMM}{XXX}
    """
    output_dir = get_po_output_dir()
    max_serial = 0
    yy = year_str[-2:] # 2026 -> 26
    
    # 정규식 패턴 설정
    if prefix_type == "MF":
        # MF260129-003 -> MF{yy}\d{4}-(\d{3})
        pattern = re.compile(rf"MF{yy}\d{{4}}-(\d{{3}})")
    else:
        # TO2601003 -> TO{yy}\d{2}(\d{3}) (월2자리+번호3자리)
        # 요청: TO + YYMM + XXX -> TO2601003
        pattern = re.compile(rf"TO{yy}\d{{2}}(\d{{3}})")

    if output_dir.exists():
        for file_path in output_dir.glob("Purchase Order_*.xlsx"):
            # 파일 내용이 아닌 파일명 등에서 찾기는 어려울 수 있음 (파일명 규칙이 지정되지 않음)
            # 파일명 규칙을 가정: Purchase Order_{MFNo}_{TONo}.xlsx
            match = pattern.search(file_path.name)
            if match:
                try:
                    serial = int(match.group(1))
                    if serial > max_serial:
                        max_serial = serial
                except ValueError:
                    continue
    
    next_serial = max_serial + 1
    return f"{next_serial:03d}"

def get_next_commercial_invoice_serial():
    today_ymd = datetime.now().strftime('%Y%m%d')
    context = today_ymd # 20260201
    
    conn = get_conn()
    cur = conn.cursor()
    
    # DB에서 마지막 번호 조회
    cur.execute("""
        SELECT last_serial FROM serial_counters 
        WHERE category = ? AND context = ?
    """, ('COMMERCIAL_INVOICE', context))
    row = cur.fetchone()
    
    if row:
        last_serial = row[0]
        next_serial = last_serial + 1
    else:
        # 해당 날짜 첫 생성
        next_serial = 1
        
    conn.close()
    
    # 001, 002 형식
    return f"{next_serial:03d}"

def generate_commercial_invoice(delivery_items, is_repair_return=False, secondary_packaging="", manual_invoice_no=None):
    """
    납품 인보이스 엑셀 생성
    :param delivery_items: 납품 항목 리스트 (Dict)
    :param is_repair_return: 수리 후 재수출 여부 (True/False)
    :param secondary_packaging: 포장/팔레트 정보 (B58 셀용)
    :param manual_invoice_no: 수동 입력한 인보이스 번호 (None이면 자동 생성)
    """
    if not delivery_items:
        return "선택된 납품 항목이 없습니다."

    try:
        # 1. 템플릿 로드 (납품 인보이스.xlsx)
        template_path = resource_path("app/templete/납품 인보이스.xlsx")
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active

        # 2. 공통 정보 설정
        now = datetime.now()
        date_str = now.strftime("%d %b, %Y") # 29 Jan, 2026
        
        # 인보이스 번호/날짜 설정
        today_ymd = now.strftime('%Y%m%d') # Default context
        seq_str = None
        
        if manual_invoice_no:
            invoice_no = manual_invoice_no
            # DB 업데이트를 위해 번호 파싱 (예: KI20260201-001 -> context=20260201, seq=1)
            # 형식이 다르면 업데이트 건너뜀
            match = re.search(r'(\d{8})[-_](\d+)', invoice_no)
            if match:
                today_ymd = match.group(1) # Override context from input
                seq_str = match.group(2)
        else:
            seq_str = get_next_commercial_invoice_serial()
            invoice_no = f"KI{today_ymd}-{seq_str}"
        
        # 사용자에게 번호 수정 기회 제공? (일관성을 위해 제공)
        # 여기서 바로 확정하지 않고, 마지막에 저장.
        # 일단 자동 생성값 사용. (UI에서 팝업 띄우는게 좋으나, 제너레이터 함수 내부에서는 어려움. 
        #  PO처럼 외부에서 번호 받거나, 여기서 임시 생성 후 저장 시 확정)
        #  일단 자동 생성.
        invoice_no = f"KI{today_ymd}-{seq_str}"
        
        # 셀 매핑
        ws['AO5'] = date_str
        ws['AO8'] = invoice_no
        
        # 3. PO 번호 (AO9) - 한 줄에 3개씩
        # delivery_items에서 order_id 식별 후 DB 조회하여 정확한 PO 번호(project_name/order_name) 가져오기
        order_ids = set()
        for item in delivery_items:
            if item.get('order_id'):
                order_ids.add(item['order_id'])
        
        po_numbers = []
        if order_ids:
            conn = get_conn()
            cur = conn.cursor()
            placeholders = ','.join('?' for _ in order_ids)
            # orders 테이블의 order_name이 PO번호라고 가정. (또는 project_name?)
            # 통상 order_name을 사용.
            cur.execute(f"SELECT order_no FROM orders WHERE id IN ({placeholders})", list(order_ids))
            po_numbers = [row[0] for row in cur.fetchall() if row[0]]
            conn.close()
            
        # PO 번호 포맷팅 (3개마다 줄바꿈)
        po_text = ""
        chunk_size = 3
        sorted_pos = sorted(po_numbers)
        chunks = [sorted_pos[i:i + chunk_size] for i in range(0, len(sorted_pos), chunk_size)]
        po_lines = [", ".join(chunk) for chunk in chunks]
        po_text = "\n".join(po_lines)
        
        ws['AO9'] = po_text
        if len(chunks) > 1:
            ws['AO9'].alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')

        # 4. 거래 구분 (일반 vs 수리)
        # 일반: AB13="SOLD", AB33="", AK33="X", AT33="", AG36="EXW"
        # 수리: AB13="수리후 재수출", AB14="NO COMMERCIAL VALUE" (Red/Bold), AB33="X", AK33="", AT33="", AG36="DDP"
        
        red_font = Font(color="FF0000", bold=True, name="Arial", size=10) # 폰트 사이즈는 템플릿 따름
        bold_font = Font(bold=True, name="Arial", size=10) # 템플릿 기본 확인 필요
        
        if is_repair_return:
            ws['AB13'] = "수리후 재수출"
            ws['AB14'] = "NO COMMERCIAL VALUE"
            ws['AB14'].font = red_font
            
            ws['AB33'] = "X" # Check Left
            ws['AK33'] = ""  # Check Right
            ws['AT33'] = ""  # Check Right 2?
            ws['AG36'] = "DDP"
        else:
            ws['AB13'] = "SOLD"
            ws['AB14'] = "" # Clear in case template has something
            
            ws['AB33'] = ""
            ws['AK33'] = "X"
            ws['AT33'] = ""
            ws['AG36'] = "EXW"

        # 5. 품목 리스트 (B43 ~ )
        # delivery_items 루프
        start_row = 43
        current_row = start_row
        
        # 파일명 생성을 위한 집계
        file_name_parts = []
        
        # 품목 정렬? (이름순?)
        sorted_items = sorted(delivery_items, key=lambda x: x.get('product_name', ''))
        
        # 동일 품목(이름+단가 등) 합치기? 인보이스는 보통 시리얼별로 안 나누고 품목별 합산함.
        # User request: "B43부터는 각 주문별로 한 행에 입력을 시작하며" -> "주문별"이 아니라 "품목별"일 듯.
        # "1번 제품명약어 xxEA, 2번 제품명약어 xxEA.xlsx" 파일명 규칙을 보면 품목별 합산이 맞음.
        
        # Group by product_name, item_code, unit_price
        grouped_items = {}
        for item in sorted_items:
            key = (item.get('product_name'), item.get('item_code'), item.get('unit_price'), 
                   item.get('items_per_box', 1), item.get('box_weight', 0.0))
            if key not in grouped_items:
                grouped_items[key] = {
                    'qty': 0,
                    'box_qty': 0.0, # 계산 필요
                    'item': item,
                    'total_weight': 0.0
                }
            grouped_items[key]['qty'] += item.get('qty', 0)
            
            # 박스 수량 계산: 
            # items_per_box가 있으면, ceil(qty / items_per_box)
            # 하지만 개별 아이템 루프이므로, 합산 후 ceil할지, 개별 ceil할지 중요.
            # 보통 합산 Qty / items_per_box -> ceil.
        
        item_idx = 1
        for key, data in grouped_items.items():
            p_name, p_code, u_price, items_per_box, box_weight = key
            qty = data['qty']
            
            # 박스 수량 계산 (올림)
            box_qty = math.ceil(qty / items_per_box) if items_per_box else 0
            
            # 단위 중량 (Unit Weight) 추정
            # DB box_weight는 박스 중량(Gross?). 
            # 인보이스에는 보통 Unit Net Weight나 Unit Gross Weight를 적음.
            # User: "I43에는 제품 1개당 중량"
            # Logic: box_weight / items_per_box (단순 평균)
        # 5. 품목 리스트 (B43 ~ )
        start_row = 43
        current_row = start_row
        grouped_items = {}
        
        # 항목 그룹화 및 정렬
        sorted_items = sorted(delivery_items, key=lambda x: x.get('item_code', ''))
        
        for item in sorted_items:
            key = (item.get('product_name'), item.get('item_code'), item.get('unit_price'), 
                   item.get('items_per_box', 1), item.get('box_weight', 0.0))
            if key not in grouped_items:
                grouped_items[key] = {'qty': 0, 'box_qty': 0.0, 'item': item}
            
            qty_val = item.get('quantity', 0)
            grouped_items[key]['qty'] += qty_val

        total_box_qty = 0
        total_net_weight = 0.0
        file_name_parts = []
        
        for key, data in grouped_items.items():
            p_name, p_code, u_price, items_per_box, box_weight = key
            qty = data['qty']
            
            # Box Qty 계산
            box_qty = math.ceil(qty / items_per_box) if items_per_box else 0
            total_box_qty += box_qty
            
            # Unit Net Weight 계산
            unit_weight = box_weight / items_per_box if items_per_box else 0
            
            # Total Net Weight 누적
            total_net_weight += unit_weight * qty
            
            # B43: Box Qty
            ws.cell(row=current_row, column=2, value=box_qty) 
            # F43: Qty
            ws.cell(row=current_row, column=6, value=qty)
            # I43: Unit Net Weight
            ws.cell(row=current_row, column=9, value=unit_weight).number_format = '0.00'
            # M43: PCS
            ws.cell(row=current_row, column=13, value="PCS")
            # Q43: Product Name
            ws.cell(row=current_row, column=17, value=p_name)
            # AH43: Part No
            ws.cell(row=current_row, column=34, value=p_code)
            # AM43: KR
            ws.cell(row=current_row, column=39, value="KR")
            # AQ43: Unit Price
            ws.cell(row=current_row, column=43, value=u_price).number_format = '#,##0.00'
            
            # 파일명 생성: 약어 사용
            abbr = ABBR_MAP.get(p_code)
            if not abbr:
                abbr = ABBR_MAP.get(p_name, p_name)
            
            safe_abbr = re.sub(r'[\\/*?:"<>|]', "", abbr).strip()
            file_name_parts.append(f"{safe_abbr} {qty}EA")
            
            current_row += 1

        # Footer Values
        # B58: Packing Info
        if secondary_packaging and "PLT" in str(secondary_packaging).upper():
            match = re.search(r'(\d+)\s*PLT', str(secondary_packaging), re.IGNORECASE)
            if match:
                 ws['B58'] = f"{match.group(1)}PLT"
            else:
                 ws['B58'] = secondary_packaging 
        else:
            ws['B58'] = total_box_qty
            
        # I58: Total Net Weight
        ws['I58'] = total_net_weight
        ws['I58'].number_format = '0.00'
        
        # Q58: Gross Weight (I58 + 1%)
        ws['Q58'] = total_net_weight * 1.01
        ws['Q58'].number_format = '0.00'
        
        # 6. 파일 저장
        output_dir = get_output_dir("Commercial Invoice")
        
        contents_str = ", ".join(file_name_parts)
        if len(contents_str) > 50: contents_str = contents_str[:50] + "..."
        
        filename = f"{invoice_no}_{contents_str}.xlsx"
        save_path = output_dir / filename
        
        wb.save(save_path)
        
        # 7. Serial Update
        if seq_str:
            try:
                update_db_serial('COMMERCIAL_INVOICE', today_ymd, int(seq_str))
            except:
                pass # 숫자가 아니거나 오류 시 무시
        
        return str(save_path)

    except Exception as e:
        return f"인보이스 생성 실패: {e}"


def generate_order_acknowledgement(order_id: int, oa_serial: str, template_path: str):
    """OA 생성 (기존 유지)"""
    # 1. 데이터 조회
    order_sql = "SELECT order_dt, order_no, req_due, final_due FROM orders WHERE id = ?"
    order_rows = query_all(order_sql, (order_id,))
    if not order_rows: raise ValueError("주문을 찾을 수 없습니다.")
    order_dt, order_no, req_due, final_due = order_rows[0]
    
    items_sql = "SELECT item_code, rev, product_name, qty, unit_price_cents FROM order_items WHERE order_id = ? ORDER BY id ASC"
    items = query_all(items_sql, (order_id,))
    
    # 2. 엑셀 템플릿 로드
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"템플릿 파일을 찾을 수 없습니다: {template_path}")
        
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active 
    
    # 3. 헤더
    today = datetime.now()
    today_str = today.strftime("%Y%m%d")
    
    oa_number = f"OA{today_str}-{oa_serial}"
    ws["AD5"] = oa_number
    
    # 오늘 날짜 포맷 (28-Jan-26)
    today_fmt = f"{today.day}-{['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'][today.month-1]}-{today.year % 100}"
    ws["AD6"] = today_fmt
    ws["J16"] = order_no
    
    def format_date_en(dt_str):
        if not dt_str: return ""
        try:
            dt = datetime.strptime(dt_str, "%Y-%m-%d")
            months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
            return f"{dt.day}-{months[dt.month-1]}-{dt.year % 100}"
        except: return dt_str

    ws["J17"] = format_date_en(order_dt)
    ws["J18"] = format_date_en(req_due)
    
    # 4. 품목
    start_row = 24
    for i, item in enumerate(items):
        row = start_row + i
        item_code, rev, prod_name, qty, unit_price_cents = item
        ws[f"B{row}"] = (i + 1) * 10
        ws[f"F{row}"] = item_code
        ws[f"K{row}"] = rev if rev else ""
        ws[f"M{row}"] = prod_name
        ws[f"AC{row}"] = qty
        ws[f"AE{row}"] = (unit_price_cents or 0) / 100
        
    # 5. 저장
    file_name = f"Order Acknowledgement_{oa_number}_{order_no}.xlsx"
    save_path = get_oa_output_dir() / file_name
    
    # [NEW] 도장 이미지 삽입
    # insert_transparent_stamp(ws)
    
    wb.save(save_path)
    
    # DB 일련번호 업데이트
    try:
        serial_int = int(oa_serial)
        update_db_serial("OA", year_str, serial_int)
    except:
        pass
        
    return str(save_path)


def generate_purchase_order(purchase_id: int, purchase_no: str, mf_serial: str, to_serial: str, template_path: str):
    """
    발주서(Purchase Order) 생성
    mf_serial: XXX (3자리 숫자 문자열)
    to_serial: XXX (3자리 숫자 문자열)
    """
    
    # 1. 데이터 조회
    # 발주 정보
    # purchases 테이블: created_at, delivery_date (납기?) 등 확인 필요.
    # purchase_widget.py -> get_purchase_with_items 등 참고
    # 여기서는 query_all로 직접 조회
    
    # purchases 테이블 구조: id, title, created_at, updated_at...
    # (주의: 발주번호 같은 필드가 명확치 않음, 보통 title이나 id 사용, 여기선 새로 생성하는 번호 사용)
    
    # 발주 품목
    # purchase_items: id, purchase_id, item_code, qty, unit_price_cents, ...
    # product_master 조인하여 rev, name 가져오기
    items_sql = """
        SELECT pi.item_code, pm.rev, pm.product_name, pi.qty, pi.unit_price_cents
        FROM purchase_items pi
        LEFT JOIN product_master pm ON pi.item_code = pm.item_code
        WHERE pi.purchase_id = ?
        ORDER BY pi.id ASC
    """
    items = query_all(items_sql, (purchase_id,))
    
    # 연결된 주문 정보 (product_name 약어 변환용 매핑)
    abbr_map = {
        "B10000805055": "MOD 29",
        "B10000850333": "38-101A",
        "B10000850460": "32-460",
        "B10000850490": "32-450",
        "B10000851180": "CRPS",
        "B10000852308": "MOD 781",
        "B10000852323": "MOD 781B"
    }

    # 연결된 주문 찾기
    # purchase_order_links 테이블: purchase_id, order_id
    # orders 테이블: final_due (최종 납기일)
    # order_items: product_name (주문의 제품명 -> 약어 변환 대상)
    # 다만 어떤 품목을 만들기 위한 발주인지 연결 고리가 명확해야 함.
    # 보통 link는 발주(Buy)가 어떤 주문(Sell)을 위해 필요한지 매핑.
    # 여러 주문이 섞여 있을 수 있음.
    
    links_sql = """
        SELECT o.final_due, oi.product_name, pol.qty_needed
        FROM purchase_order_links pol
        JOIN orders o ON pol.order_id = o.id
        JOIN order_items oi ON pol.order_item_id = oi.id
        WHERE pol.purchase_id = ?
        ORDER BY o.final_due ASC
    """
    # 주의: purchase_order_links에 order_item_id가 있는지 확인 필요.
    # db.py 스키마 확인을 못했으나, 보통은 상세 매핑이 필요함.
    # 만약 order_id만 있다면, 해당 주문의 메인 제품을 찾아야 함.
    # 스키마를 모르는 상태에서 가장 안전한 건 order_id로 orders 조인하고, 
    # orders <- order_items (보통 1:N 이지만, 1:1 메인 장비 가정 시)
    # 일단 위 쿼리 시도해보고 컬럼 에러 나면 수정.
    # -> 기존 대화 로그나 파일 뷰에서 확인된 바 없음. 
    # -> 안전하게 order_id로만 조회하고, order_items에서 Representative Item을 가져오거나 해야 함.
    # -> 보통 주문 1개당 장비 1대면 order_items 첫번째꺼 가져옴.
    
    # 대체 쿼리 (Purchase -> Order Link -> Order -> Order Items)
    linked_info = []
    
    # [추가] 필요한 모듈 임포트 (함수 상단으로 이동)
    from copy import copy
    from openpyxl.styles import Side, Border

    # 2. 엑셀 로드
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"템플릿 파일을 찾을 수 없습니다: {template_path}")
    
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    
    # 3. 헤더 입력
    today = datetime.now()
    yy = today.strftime("%y")   # 26
    ymd = today.strftime("%y%m%d") # 260129
    ym = today.strftime("%y%m") # 2601
    
    # C3: MF + YYMMDD + - + XXX
    mf_number = f"MF{ymd}-{mf_serial}"
    ws["C3"] = mf_number
    
    # C4: 작성날짜 (YYYY-MM-DD)
    ws["C4"] = today.strftime("20%y-%m-%d")
    
    # H4: TO + YYMM + XXX -> Request: Just use Purchase No
    # to_number = f"TO{ym}{to_serial}"
    ws["H4"] = purchase_no
    
    # 4. 품목 리스트 (15행~)
    # A=No, B=Code, C=Rev, D=Desc, E=Unit, F=Qty, G=Price
    start_row = 15
    for i, item in enumerate(items):
        row = start_row + i
        item_code, rev, prod_name, qty, unit_price_cents = item
        
        ws[f"A{row}"] = i + 1
        ws[f"B{row}"] = item_code
        ws[f"C{row}"] = rev if rev else ""
        ws[f"D{row}"] = prod_name
        ws[f"E{row}"] = "EA"
        ws[f"F{row}"] = qty
        # 단가: 원(KRW) 가정, cents / 100?
        # 보통 발주 단가는 원화일 가능성 높음. 
        # 사용자가 "단가를 넣음"이라고만 함. -> 그대로 넣되 정수 포맷 추천.
        price = (unit_price_cents or 0) / 100
        ws[f"G{row}"] = price
        
        # 5. 연결 정보 (D25)
        # [제품명 약어] [수량]대 : [납기일](일본 ULVAC-PHI 도착 기준)
        # 링크 정보 조회 시도
        conn = get_conn()
        cur = conn.cursor()
        
        try:
            # -------------------------------------------------------------------------
            # Section 5: 연결 주문 정보 (Linked Order Information)
            # -------------------------------------------------------------------------
            # [수정] 주문별 상세 납기일 정보를 order_shipments에서 가져옴 (분할 납기 지원)
            # 만약 shipment 정보가 없으면 orders.final_due를 기입함.
            cur.execute("""
                SELECT 
                    COALESCE(os.due_date, o.final_due) as effective_due,
                    oi.product_name,
                    COALESCE(os.ship_qty, oi.qty) as effective_qty,
                    oi.item_code
                FROM purchase_order_links pol
                JOIN orders o ON pol.order_id = o.id
                JOIN order_items oi ON o.id = oi.order_id
                LEFT JOIN order_shipments os ON oi.id = os.order_item_id
                WHERE pol.purchase_id = ?
            """, (purchase_id,))
            
            rows = cur.fetchall()
            
            # 데이터 가공: (납기, 약어) 별 수량 집계
            from collections import defaultdict
            grouped = defaultdict(int) 
            
            for r in rows:
                f_due, p_name, q, i_code = r
                # 제품명 매핑
                abbr = abbr_map.get(i_code)
                if not abbr:
                    abbr = abbr_map.get(p_name, p_name)
                
                f_due_key = str(f_due).strip() if f_due else ""
                grouped[(f_due_key, abbr)] += q
            
            # D25부터 출력
            link_start_row = 25
            max_fixed_rows = 3  # D25, D26, D27
            sorted_keys = sorted(grouped.keys(), key=lambda x: x[0])
            
            # ✅ [Fix] 행 삽입 전 푸터 영역 높이 및 이미지 위치를 위한 높이 캡처
            # 단순 'in ws.row_dimensions' 체크 대신 넉넉한 범위를 모두 저장
            footer_heights = {}
            capture_start = link_start_row + max_fixed_rows # 28
            for r in range(capture_start, capture_start + 150):
                if r in ws.row_dimensions:
                    footer_heights[r] = ws.row_dimensions[r].height
            
            # ✅ [Fix] 헤더 영역(5~12행) 점선 테두리 문제 해결
            # 템플릿 파일 자체에 점선이 들어있는 것으로 추정되므로, 강제로 실선으로 변경
            thin_border = Side(style='thin')
            for r in range(5, 13):
                # K, L열을 포함한 주요 컬럼 체크 (F~M)
                for c in range(6, 15): 
                    cell = ws.cell(row=r, column=c)
                    if cell.has_style and cell.border:
                        b = copy(cell.border)
                        changed = False
                        # 점선(dashed, dotted 등이면 실선으로 변경)
                        # 명시적으로 'thin'이 아닌 선이 있으면 강제 수정
                        target_styles = ['dashed', 'dotted', 'mediumDashed', 'hair', 'medium']
                        if b.left and b.left.style in target_styles: 
                            b.left = thin_border
                            changed = True
                        if b.right and b.right.style in target_styles:
                            b.right = thin_border
                            changed = True
                        
                        if changed:
                            cell.border = b

            # ✅ 이미지 이동 헬퍼 (행 삽입 시 아래로 밀기)
            def shift_images(ws, start_row, shift_count=1):
                threshold_idx = start_row - 1
                if hasattr(ws, '_images'):
                    for img in ws._images:
                        if hasattr(img.anchor, '_from') and img.anchor._from.row >= threshold_idx:
                            img.anchor._from.row += shift_count
                        if hasattr(img.anchor, 'to') and img.anchor.to.row >= threshold_idx:
                            img.anchor.to.row += shift_count

            # ✅ 스타일 복사 헬퍼
            def copy_row_style(ws, src_row_idx, tgt_row_idx):
                """src_row의 스타일을 tgt_row로 복사 (병합 포함)"""
                src_row = ws[src_row_idx]
                tgt_row = ws[tgt_row_idx]
                for src_cell, tgt_cell in zip(src_row, tgt_row):
                    if src_cell.has_style:
                        tgt_cell.font = copy(src_cell.font)
                        tgt_cell.border = copy(src_cell.border)
                        tgt_cell.fill = copy(src_cell.fill)
                        tgt_cell.number_format = src_cell.number_format
                        tgt_cell.protection = copy(src_cell.protection)
                        tgt_cell.alignment = copy(src_cell.alignment)
                
                # 병합 복사
                source_merges = [m for m in ws.merged_cells.ranges if m.min_row == src_row_idx and m.max_row == src_row_idx]
                for merge in source_merges:
                    ws.merge_cells(start_row=tgt_row_idx, start_column=merge.min_col,
                                   end_row=tgt_row_idx, end_column=merge.max_col)

            current_link_row = link_start_row
            total_inserted_rows = 0
            
            for idx, (f_due, abbr) in enumerate(sorted_keys):
                qty = grouped[(f_due, abbr)]
                
                # ✅ 동적 행 삽입 (3개 초과 시)
                if idx >= max_fixed_rows:
                    ws.insert_rows(current_link_row)
                    total_inserted_rows += 1
                    # [중요] 스타일은 푸터가 아닌 '데이터 행(25행)'에서 복사해와야 깨끗함
                    copy_row_style(ws, 25, current_link_row)
                    
                    # ✅ [Fix] 복사된 스타일에서 불필요한 가로선(Bottom Border) 제거
                    # 25행이 박스 형태라 테두리가 딸려오는 문제 해결
                    no_border_side = Side(border_style=None)
                    for col_idx in range(1, 15): # A~N 까지 넉넉하게 체크 (J, K열 포함)
                        cell = ws.cell(row=current_link_row, column=col_idx)
                        if cell.has_style and cell.border:
                            new_border = copy(cell.border)
                            new_border.top = no_border_side
                            new_border.bottom = no_border_side
                            cell.border = new_border

                    shift_images(ws, current_link_row, 1)
                
                ws.row_dimensions[current_link_row].height = 20
                
                # 날짜 출력 포맷
                date_str = ""
                if f_due:
                    try:
                        # 다양한 날짜 형식 대응 (YYYY-MM-DD or YYYY-MM-DD HH:MM:SS)
                        clean_due = f_due.split(' ')[0]
                        dt = datetime.strptime(clean_due, "%Y-%m-%d")
                        date_str = f"{dt.year}년 {dt.month}월 {dt.day}일"
                    except:
                        date_str = f_due
                else:
                    date_str = "(납기미정)"
                    
                text = f"{idx+1}) {abbr} {qty}대 : {date_str}(일본 ULVAC-PHI 도착 기준)"
                ws[f"D{current_link_row}"] = text
                current_link_row += 1

            # ✅ [Fix] 푸터 영역 행 높이 원상 복구 (기존 로직 + 강제 보정)
            if total_inserted_rows > 0:
                for r_orig, h in footer_heights.items():
                    ws.row_dimensions[r_orig + total_inserted_rows].height = h
            
            # ✅ [Robust Fix] 로고 행(38행 등)이 찌그러지는 문제 절대 방지
            # "권성관 대표" 또는 "세금계산서"가 있는 행을 찾아서 그 다음 행(로고 위치)의 높이를 강제 설정
            footer_scan_start = 35 + total_inserted_rows # 대략적인 위치부터 검색
            for r in range(footer_scan_start, footer_scan_start + 20):
                cell_val = str(ws[f"D{r}"].value or "")
                if "권성관" in cell_val or "baguman" in cell_val:
                    # 권성관 대표 행 발견 (예: 36행)
                    # 사용자 피드백 기반 높이 보정
                    # 37행(N+1): 20
                    ws.row_dimensions[r + 1].height = 20
                    # 38행(N+2): 25
                    ws.row_dimensions[r + 2].height = 25
                    # 39행(N+3): 25
                    ws.row_dimensions[r + 3].height = 25
                    break

            # ✅ [Fix] Linked Order 영역(25행~) 및 푸터 점선 테두리 제거 (25행~현재+20행 스캔)
            thin_border = Side(style='thin')
            scan_end_row = current_link_row + 20
            target_styles = ['dashed', 'dotted', 'mediumDashed', 'hair', 'medium']
            
            for r in range(25, scan_end_row):
                # F(6) ~ N(14) 열 스캔
                for c in range(6, 15): 
                    cell = ws.cell(row=r, column=c)
                    if cell.has_style and cell.border:
                        b = copy(cell.border)
                        changed = False
                        
                        if b.left and b.left.style in target_styles: 
                            b.left = thin_border
                            changed = True
                        if b.right and b.right.style in target_styles:
                            b.right = thin_border
                            changed = True
                        
                        if changed:
                            cell.border = b
                            
            # ✅ 인쇄 영역 업데이트
            if ws.print_area and total_inserted_rows > 0:
                try:
                    p_area = str(ws.print_area)
                    if ':' in p_area:
                        parts = p_area.split(':')
                        match = re.match(r"([A-Z]+)(\d+)", parts[1])
                        if match:
                            ws.print_area = f"{parts[0]}:{match.group(1)}{int(match.group(2)) + total_inserted_rows}"
                except:
                    pass
            
            # ✅ [Fix] 푸터 잔여 세로선(Artifact) 제거 로직 강화
            # D38, D39, D40 등에서 발생하는 의도치 않은 우측 세로선 제거
            # 조건부(키워드 체크) 삭제가 아니라 범위 내 무조건 삭제로 변경하여 확실하게 제거
            footer_scan_start = current_link_row + 1
            clear_border_side = Side(border_style=None) 
            
            for r in range(footer_scan_start, footer_scan_start + 30):
                d_cell = ws[f"D{r}"]
                # 내용이 있든 없든 테두리 삭제
                if d_cell.has_style:
                    b = copy(d_cell.border)
                    b.right = clear_border_side
                    d_cell.border = b
                
                # E셀의 좌측 테두리도 삭제
                e_cell = ws[f"E{r}"]
                if e_cell.has_style and e_cell.border:
                    eb = copy(e_cell.border)
                    eb.left = clear_border_side
                    e_cell.border = eb                
        except Exception as e:
            print(f"Link Info Error: {e}")
        finally:
            conn.close()

    # 6. 저장
    # 파일명 규칙: 
    # 발주번호_발주서_H5셀의 업체명_YYMMDD_제품명약어+스페이스+X대(발주 품목이 다수일 경우에는 콤마(,) 스페이스 후에 제품명약어+스페이스+X대)
    
    # H5 (업체명)
    vendor_name = ws["H5"].value or "Unknown"
    vendor_name = str(vendor_name).replace("주식회사", "").strip()
    
    # 제품명약어+스페이스+X대
    # 사용자 요청: 파일명에는 "주문정보와 관계없이 발주내용의 품목의 수량"을 넣어야 함.
    # 즉, linked_info가 아니라 items(발주 품목)를 기준으로 집계.
    
    summary_map = defaultdict(int) 
    for item in items:
        # item structure: item_code, rev, prod_name, qty, unit_price_cents
        i_code = item[0]
        p_name = item[2]
        i_qty = item[3]
        
        # 약어 매핑
        abbr = abbr_map.get(i_code)
        if not abbr:
            abbr = abbr_map.get(p_name, p_name)
            
        summary_map[abbr] += i_qty
        
    summary_list = []
    # 정렬 (이름순)
    for abbr in sorted(summary_map.keys()):
        total_q = summary_map[abbr]
        summary_list.append(f"{abbr} {total_q}대")
        
    if not summary_list:
        items_str = "품목없음"
    else:
        items_str = ", ".join(summary_list)
        
    # 파일명 생성 및 특수문자 제거
    def sanitize_filename(name):
        return re.sub(r'[\\/*?:"<>|]', "", name)

    file_name_base = f"{purchase_no}_발주서_{vendor_name}_{ymd}_{items_str}"
    file_name = sanitize_filename(file_name_base) + ".xlsx"
    
    save_path = get_po_output_dir() / file_name

    # [NEW] 도장 이미지 삽입 (위치 강제 지정: E10)
    # insert_transparent_stamp(ws, fixed_anchor="E10")
    
    wb.save(save_path)

    return str(save_path)
    
def get_invoice_output_dir():
    return get_output_dir("Invoice")

def get_next_invoice_serial(date_str: str) -> str:
    """청구서 일련번호 (DB 기반)"""
    # context: '20260201'
    next_val = get_next_db_serial("INVOICE", date_str)
    return f"{next_val:03d}"


def generate_invoice(delivery_id: int, template_path: str, invoice_serial: str = None):
    """
    청구서(Invoice) 생성
    한 납품에 여러 주문이 섞여있을 수 있으며, 주문 당 하나의 파일 생성.
    invoice_serial: 사용자가 지정한 일련번호 (None이면 자동 생성)
    """
    
    # 1. 납품에 포함된 주문 ID 목록 조회
    conn = get_conn()
    cur = conn.cursor()
    
    try:
        # 1. 납품에 연결된 주문 ID 목록 수집 (머징 테이블 + 개별 항목)
        cur.execute("""
            SELECT DISTINCT order_id FROM delivery_order_links WHERE delivery_id = ?
            UNION
            SELECT DISTINCT order_id FROM delivery_items WHERE delivery_id = ? AND order_id IS NOT NULL
        """, (delivery_id, delivery_id))
        
        order_ids = [row[0] for row in cur.fetchall()]
        
        generated_files = []
        
        today = datetime.now()
        ymd_full = today.strftime("%Y%m%d") # 20260201
        
        # 다음 달 15일 계산
        if today.month == 12:
            next_month = today.replace(year=today.year+1, month=1, day=15)
        else:
            next_month = today.replace(month=today.month+1, day=15)
        payment_due_str = next_month.strftime("%Y-%m-%d")
        
        # 약어 맵
        abbr_map = {
            "B10000805055": "MOD 29",
            "B10000850333": "38-101A",
            "B10000850460": "32-460",
            "B10000850490": "32-450",
            "B10000851180": "CRPS",
            "B10000852308": "MOD 781",
            "B10000852323": "MOD 781B"
        }

        # 일련번호 초기값 설정 (정수형)
        current_serial_int = 1
        if invoice_serial:
             try:
                 current_serial_int = int(invoice_serial)
             except:
                 current_serial_int = get_next_db_serial("INVOICE", ymd_full)
        else:
             current_serial_int = get_next_db_serial("INVOICE", ymd_full)
        
        for order_id in order_ids:
            # 2. 주문 정보 조회
            order_row = query_all("SELECT order_no FROM orders WHERE id = ?", (order_id,))
            if not order_row: continue
            order_no = order_row[0][0]
            
            # 3. 해당 납품+주문의 품목 조회
            if len(order_ids) == 1:
                # 주문이 단 하나만 연결된 경우, order_id가 NULL인 항목(재고/수리품 등)도 해당 주문 인보이스에 포함
                items_sql = """
                    SELECT di.item_code, pm.rev, di.product_name, SUM(di.qty), oi.unit_price_cents
                    FROM delivery_items di
                    LEFT JOIN product_master pm ON di.item_code = pm.item_code
                    LEFT JOIN order_items oi ON oi.order_id = ? AND di.item_code = oi.item_code
                    WHERE di.delivery_id = ? AND (di.order_id = ? OR di.order_id IS NULL)
                    GROUP BY di.item_code, di.product_name
                """
                items = query_all(items_sql, (order_id, delivery_id, order_id))
            else:
                # 주문이 여러 개인 경우, 명시적으로 해당 order_id가 지정된 항목만 포함
                items_sql = """
                    SELECT di.item_code, pm.rev, di.product_name, SUM(di.qty), oi.unit_price_cents
                    FROM delivery_items di
                    LEFT JOIN product_master pm ON di.item_code = pm.item_code
                    LEFT JOIN order_items oi ON di.order_id = oi.order_id AND di.item_code = oi.item_code
                    WHERE di.delivery_id = ? AND di.order_id = ?
                    GROUP BY di.item_code, di.product_name
                """
                items = query_all(items_sql, (delivery_id, order_id))

            if not items: continue
            
            # 4. 엑셀 생성
            if not os.path.exists(template_path):
                raise FileNotFoundError(f"템플릿 없음: {template_path}")
                
            wb = openpyxl.load_workbook(template_path)
            ws = wb.active
            
            # 일련번호 및 청구서 번호
            serial_str = f"{current_serial_int:03d}"
            invoice_no = f"KI{ymd_full}-{serial_str}"
            
            # 헤더
            # Z5: Invoice No
            ws["Z5"] = invoice_no
            
            # 날짜 포맷 (DD-Mon-YY)
            months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
            
            # Z6: Date
            ws["Z6"] = f"{today.day}-{months[today.month-1]}-{today.year % 100}"
            
            # J16: Order No
            ws["J16"] = order_no
            
            # J20: Payment Due
            ws["J20"] = f"{next_month.day}-{months[next_month.month-1]}-{next_month.year % 100}"
            
            # [Style Fix] 우측 외곽선 복구
            # 머지된 셀의 경우 특정 위치 셀만 설정하면 테두리가 누락될 수 있으므로, 
            # 헤더 구역(5~20행)의 우측 끝을 포함하는 모든 전방위 셀(AI~AJ열)에 대해 강제 설정.
            thick_side = Side(style='medium')
            for row_idx in [5, 6, 16, 17, 18, 19, 20]:
                # AJ5, AJ6, AJ16... 뿐만 아니라 AI열 등 결합된 모든 셀에 대해 우측 테두리 적용
                for col_letter in ["AI", "AJ"]:
                    cell = ws[f"{col_letter}{row_idx}"]
                    cell.border = Border(
                        left=cell.border.left,
                        right=thick_side,
                        top=cell.border.top,
                        bottom=cell.border.bottom
                    )
            
            # 품목 (23행~)
            start_row = 23
            
            # 파일명용 요약
            filename_items = []
            
            for i, item in enumerate(items):
                r = start_row + i
                i_code, rev, p_name, qty, price_cents = item
                
                ws[f"D{r}"] = i_code
                ws[f"I{r}"] = rev if rev else ""
                ws[f"K{r}"] = p_name
                ws[f"Z{r}"] = qty
                
                # 단가
                price = (price_cents or 0) / 100
                ws[f"AB{r}"] = price
                
                # 파일명용 약어 처리
                abbr = abbr_map.get(i_code)
                if not abbr: abbr = abbr_map.get(p_name, p_name)
                filename_items.append(f"{abbr} {qty}EA")
                
            # 저장
            # 파일명: ULVAC-PHI_Invoice_청구서번호_품목명약어 xxEA(주문번호).xlsx
            items_str = ", ".join(filename_items)
            
            def sanitize_filename(name):
                return re.sub(r'[\\/*?:"<>|]', "", name)
            
            file_name_base = f"ULVAC-PHI_Invoice_{invoice_no}_{items_str}({order_no})"
            file_name = sanitize_filename(file_name_base) + ".xlsx"
            
            # [NEW] 도장 이미지 삽입
            # insert_transparent_stamp(ws)

            save_path = get_invoice_output_dir() / file_name
            wb.save(save_path)
            
            generated_files.append(str(save_path))

            # DB 업데이트 및 다음 번호 증가
            update_db_serial("INVOICE", ymd_full, current_serial_int)
            current_serial_int += 1
        
    finally:
        conn.close()

    return generated_files
