# app/ui/bom_widget.py
from PySide6 import QtWidgets, QtCore, QtGui
from PySide6.QtCore import Qt
from ..db import (
    get_conn, get_all_product_master, add_bom_item, delete_bom_item, 
    update_bom_item, get_bom_tree, check_circular_dependency,
    add_or_update_product_master, search_product_master, update_product_master
)
from .product_master_widget import WhereUsedDialog, ProductMasterWidget
import openpyxl

class BomWidget(QtWidgets.QWidget):
    def __init__(self, parent=None, settings=None):
        super().__init__(parent)
        self.settings = settings
        
        # ✅ 저장된 필터 로드 (없으면 기본 3가지 모두 선택)
        if self.settings:
            saved_filters = self.settings.value("filters/bom_source_types", ["PRODUCT", "MODULE", "PART"])
            self.selected_types = set(saved_filters)
        else:
            self.selected_types = {"PRODUCT", "MODULE", "PART"}

        # ✅ 저장된 '판매가능 품목만 보기' 설정 로드 (기본값: True)
        if self.settings:
            # 0이면 False, 그 외(1, 2, true 등)면 True. value()가 int나 bool을 리턴할 수 있음.
            # safe conversion
            val = self.settings.value("filters/bom_sellable_only", True)
            self.sellable_only = (str(val).lower() != 'false' and str(val) != '0')
        else:
            self.sellable_only = True
            
        self.setup_ui()
        self.load_product_list()

    def load_data_if_needed(self):
        """탭 전환 시 호출되는 메서드"""
        # 필요하다면 여기서 데이터 리로딩 (지금은 load_product_list가 init에서 불림)
        # 하지만 BOM 트리는 current_parent에 의존하므로 리로드 불필요할 수도 있음.
        # 품목 목록(소스)은 업데이트가 필요할 수 있음.
        self.load_product_list()

    def setup_ui(self):
        main_layout = QtWidgets.QHBoxLayout(self)

        # --- Left Pane: Product Master List (Source) ---
        left_layout = QtWidgets.QVBoxLayout()
        
        lbl_source = QtWidgets.QLabel("품목 목록 (소스)")
        lbl_source.setStyleSheet("font-weight: bold; font-size: 14px;")
        
        self.search_bar = QtWidgets.QLineEdit()
        self.search_bar.setPlaceholderText("품목 검색 (코드/명칭)...")
        self.search_bar.textChanged.connect(self.filter_product_list)

        self.product_list = QtWidgets.QListWidget()
        self.product_list.setDragEnabled(True) # 드래그 가능
        self.product_list.setSelectionMode(QtWidgets.QAbstractItemView.SingleSelection)
        self.product_list.setToolTip("이 목록에서 품목을 드래그하여 오른쪽 BOM 트리에 추가하세요.")
        self.product_list.setContextMenuPolicy(Qt.CustomContextMenu)
        self.product_list.customContextMenuRequested.connect(self.show_product_list_context_menu)

        # ✅ 유형 필터 버튼 (ToolButton + Menu)
        filter_layout = QtWidgets.QHBoxLayout()
        
        self.btn_type_filter = QtWidgets.QToolButton()
        self.btn_type_filter.setText("유형 필터")
        self.btn_type_filter.setPopupMode(QtWidgets.QToolButton.InstantPopup)
        
        self.type_filter_menu = QtWidgets.QMenu(self.btn_type_filter)
        self.filter_options = [("완제품 (PRODUCT)", "PRODUCT"), ("모듈 (MODULE)", "MODULE"), ("부품 (PART)", "PART")]
        
        for label, code in self.filter_options:
            action = QtWidgets.QWidgetAction(self.type_filter_menu)
            cb = QtWidgets.QCheckBox(label)
            cb.setStyleSheet("QCheckBox { padding: 5px; }")
            cb.setChecked(code in self.selected_types)
            cb.stateChanged.connect(lambda state, c=code: self.on_filter_changed(c, state))
            
            action.setDefaultWidget(cb)
            self.type_filter_menu.addAction(action)
        
        self.btn_type_filter.setMenu(self.type_filter_menu)
        self.update_filter_button_text() # 초기 텍스트 설정
        
        # ✅ 판매가능 품목만 보기 체크박스 추가
        self.chk_sellable_only = QtWidgets.QCheckBox("판매가능 품목만 보기")
        self.chk_sellable_only.setChecked(self.sellable_only)
        self.chk_sellable_only.stateChanged.connect(self.on_sellable_only_changed)

        filter_layout.addWidget(lbl_source)
        filter_layout.addStretch()
        filter_layout.addWidget(self.chk_sellable_only) # ✅ 체크박스 추가
        filter_layout.addWidget(self.btn_type_filter)

        left_layout.addLayout(filter_layout) # ✅ 레이아웃 변경
        left_layout.addWidget(self.search_bar)
        left_layout.addWidget(self.product_list)

        # --- Right Pane: BOM Tree (Target) ---
        right_layout = QtWidgets.QVBoxLayout()

        lbl_target = QtWidgets.QLabel("BOM 구성 (타겟)")
        lbl_target.setStyleSheet("font-weight: bold; font-size: 14px;")

        # Target Product Selection
        target_select_layout = QtWidgets.QHBoxLayout()
        self.lbl_current_parent = QtWidgets.QLabel("편집할 모품목을 왼쪽에서 더블클릭하세요.")
        self.lbl_current_parent.setStyleSheet("color: blue; font-weight: bold;")
        target_select_layout.addWidget(self.lbl_current_parent)
        target_select_layout.addStretch()

        # ✅ [추가] BOM 트리 검색바
        self.target_search_bar = QtWidgets.QLineEdit()
        self.target_search_bar.setPlaceholderText("BOM 구성 품목 검색 (코드/명칭)...")
        self.target_search_bar.textChanged.connect(self.filter_bom_tree)

        self.current_parent_code = None
        self.current_parent_name = None

        self.bom_tree = QtWidgets.QTreeWidget()
        # ✅ [수정] 헤더 확장: No., 자식코드, 품목명, 소요량, 단위, 유형, 비고
        self.bom_tree.setHeaderLabels(["No.", "자식 품목코드", "품목명", "소요량", "단위", "유형", "비고"])
        self.bom_tree.setAlternatingRowColors(True)
        
        # 기본 컬럼 폭 설정
        self.bom_tree.setColumnWidth(0, 50)  # No.
        self.bom_tree.setColumnWidth(1, 150) # Code
        self.bom_tree.setColumnWidth(2, 250) # Name
        self.bom_tree.setColumnWidth(3, 80)  # Qty
        self.bom_tree.setColumnWidth(4, 60)  # Unit
        self.bom_tree.setColumnWidth(5, 100) # Type (한글)
        self.bom_tree.setColumnWidth(6, 150) # Remarks

        self.bom_tree.setAcceptDrops(True) # 드롭 허용
        self.bom_tree.setDragDropMode(QtWidgets.QAbstractItemView.DropOnly)
        self.bom_tree.setSelectionMode(QtWidgets.QAbstractItemView.ExtendedSelection) # ✅ 다중 선택 활성화
        self.bom_tree.setContextMenuPolicy(Qt.CustomContextMenu)
        self.bom_tree.customContextMenuRequested.connect(self.show_context_menu)
        self.bom_tree.itemDoubleClicked.connect(self.edit_item) # ✅ edit_qty -> edit_item 변경

        # ✅ 컬럼 폭 저장/복원
        self.bom_tree.header().setSectionResizeMode(QtWidgets.QHeaderView.Interactive)
        self.bom_tree.header().sectionResized.connect(self.save_column_widths)
        if self.settings:
            self.restore_column_widths()

        # ✅ 컬럼 리사이즈 정책 적용
        from .utils import apply_table_resize_policy
        apply_table_resize_policy(self.bom_tree)
        
        # ✅ [수정] 가로 스크롤바 활성화 (utils 정책 오버라이드)
        self.bom_tree.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.bom_tree.header().setStretchLastSection(False)

        # 기능 버튼
        btn_layout = QtWidgets.QHBoxLayout()
        self.btn_add_child = QtWidgets.QPushButton("◀ 자식 추가 (선택한 품목)")
        self.btn_add_child.setToolTip("왼쪽 목록에서 선택한 품목을 현재 BOM의 자식으로 추가합니다.")
        
        # ✅ [추가] 직접 입력 버튼
        self.btn_manual_add = QtWidgets.QPushButton("직접 입력")
        self.btn_manual_add.setToolTip("목록에 없는 품목 코드를 직접 입력하여 추가합니다.")

        self.btn_remove_child = QtWidgets.QPushButton("선택 삭제")
        
        self.btn_expand = QtWidgets.QPushButton("모두 펼치기")
        self.btn_collapse = QtWidgets.QPushButton("모두 접기")
        self.btn_excel_import = QtWidgets.QPushButton("엑셀 가져오기")
        self.btn_excel_export = QtWidgets.QPushButton("엑셀 내보내기")

        self.btn_add_child.clicked.connect(self.on_btn_add_child_clicked)
        self.btn_manual_add.clicked.connect(self.on_btn_manual_add_clicked) # ✅ 연결
        self.btn_remove_child.clicked.connect(self.on_btn_remove_child_clicked)
        self.btn_expand.clicked.connect(self.bom_tree.expandAll)
        self.btn_collapse.clicked.connect(self.bom_tree.collapseAll)
        self.btn_excel_export.clicked.connect(self.export_to_excel)
        self.btn_excel_import.clicked.connect(self.import_from_excel)

        # 버튼 배치 row 1
        btn_row1 = QtWidgets.QHBoxLayout()
        btn_row1.addWidget(self.btn_add_child)
        btn_row1.addWidget(self.btn_manual_add) # ✅ 추가
        btn_row1.addWidget(self.btn_remove_child)
        btn_row1.addStretch()
        
        # 버튼 배치 row 2
        btn_row2 = QtWidgets.QHBoxLayout()
        btn_row2.addWidget(self.btn_expand)
        btn_row2.addWidget(self.btn_collapse)
        btn_row2.addStretch()
        btn_row2.addWidget(self.btn_excel_import)
        btn_row2.addWidget(self.btn_excel_export)
        
        # 도움말 라벨
        self.lbl_help = QtWidgets.QLabel("💡 팁: 목록에서 품목을 드래그하여 트리에 놓으면 쉽게 추가할 수 있습니다.")
        self.lbl_help.setStyleSheet("color: gray; font-style: italic; margin-top: 5px;")

        right_layout.addWidget(lbl_target)
        right_layout.addLayout(target_select_layout)
        right_layout.addWidget(self.target_search_bar) # ✅ 레이아웃 추가
        right_layout.addWidget(self.bom_tree)
        right_layout.addWidget(self.lbl_help) # 도움말 추가
        right_layout.addLayout(btn_row1)      # 추가/삭제 버튼
        right_layout.addLayout(btn_row2)      # 기타 버튼

        # --- Splitter ---
        left_widget = QtWidgets.QWidget()
        left_widget.setLayout(left_layout)
        right_widget = QtWidgets.QWidget()
        right_widget.setLayout(right_layout)

        splitter = QtWidgets.QSplitter(Qt.Horizontal)
        splitter.addWidget(left_widget)
        splitter.addWidget(right_widget)
        splitter.setStretchFactor(0, 1)
        splitter.setStretchFactor(1, 2)

        main_layout.addWidget(splitter)

        # Events
        self.product_list.itemDoubleClicked.connect(self.highlight_logic_check)
        
        # Override dropEvent for TreeWidget is tricky without subclassing, 
        # so we will install an event filter or subclass. 
        # For simplicity, let's subclass QTreeWidget inline or monkey patch if possible, 
        # but subclassing is cleaner. However, here I will attach an event filter.
        self.bom_tree.viewport().installEventFilter(self)


    def load_product_list(self):
        """왼쪽 목록에 모든 활성 품목 로드 (필터 적용)"""
        self.product_list.clear()
        
        # ✅ 필터 목록 구성
        target_types = list(self.selected_types)
        # 레거시 데이터 및 한글 데이터 호환
        if 'PRODUCT' in self.selected_types: 
            target_types.extend(['SELLABLE', '완제품', 'Product'])
        if 'MODULE' in self.selected_types: 
            target_types.extend(['모듈', 'Module'])
        if 'PART' in self.selected_types: 
            target_types.extend(['SUB_COMPONENT', '부품', 'Part'])
        
        if not target_types:
            target_types = ['__NO_MATCH__']

        # DB에서 필터링하여 가져옴 (판매가능 여부 포함)
        # 체크박스가 켜져있으면 include_inactive=False (활성만), 꺼져있으면 True (모두)
        include_inactive = not self.chk_sellable_only.isChecked()
        
        products = get_all_product_master(include_inactive=include_inactive, type_filter=target_types)
        self.all_products_cache = products
        
        for p in products:
            # p: id, code, rev, name, ..., item_type
            code = p[1]
            rev = p[2]
            name = p[3]
            item_type = p[10] 
            
            # 한글 표기 변환
            type_map = {
                'PRODUCT': '완제품',
                'MODULE': '모듈',
                'PART': '부품',
                'SELLABLE': '완제품(구)',
                'SUB_COMPONENT': '부품(구)' 
            }
            type_str = type_map.get(item_type, item_type)
            
            display_text = f"[{type_str}] {code} - {name}"
            if rev: display_text += f" (Rev {rev})"
            
            item = QtWidgets.QListWidgetItem(display_text)
            item.setData(Qt.UserRole, {'code': code, 'name': name, 'type': item_type})
            self.product_list.addItem(item)
            
        # ✅ 로드 후 현재 검색어 필터 재적용 (사용자 경험 유지)
        current_search = self.search_bar.text()
        if current_search:
            self.filter_product_list(current_search)
            
    def on_filter_changed(self, code, state):
        """유형 필터 변경 핸들러"""
        if state == 2:  # Checked
            self.selected_types.add(code)
        else:
            self.selected_types.discard(code)
            
        self.update_filter_button_text()
        
        if self.settings:
            self.settings.setValue("filters/bom_source_types", list(self.selected_types))
            
        self.load_product_list()

    def on_sellable_only_changed(self, state):
        """판매가능 품목만 보기 체크박스 변경 핸들러"""
        is_checked = (state == 2)
        self.sellable_only = is_checked
        
        if self.settings:
            self.settings.setValue("filters/bom_sellable_only", is_checked)
            
        self.load_product_list()

    def update_filter_button_text(self):
        count = len(self.selected_types)
        total = len(self.filter_options)
        self.btn_type_filter.setText(f"유형 필터 ({count}/{total})")

    def filter_product_list(self, text):
        search_text = text.lower()
        for i in range(self.product_list.count()):
            item = self.product_list.item(i)
            # 숨기기/보이기
            item.setHidden(search_text not in item.text().lower())

    def filter_bom_tree(self, text):
        """BOM 트리 필터링 (검색어 포함 항목 + 그 조상만 표시)"""
        search_text = text.lower()
        
        # 최상위 아이템 순회
        root = self.bom_tree.invisibleRootItem()
        for i in range(root.childCount()):
            item = root.child(i)
            self.filter_tree_item(item, search_text)

    def filter_tree_item(self, item, search_text):
        """재귀적으로 트리 아이템 필터링"""
        # 1: Code, 2: Name
        code = item.text(1).lower()
        name = item.text(2).lower()
        match = (search_text in code) or (search_text in name)
        
        # 자식들 확인
        child_match = False
        for i in range(item.childCount()):
            child = item.child(i)
            if self.filter_tree_item(child, search_text):
                child_match = True
                
        # 나 자신이 매치되거나, 자식 중에 매치되는 게 있으면 표시
        should_show = match or child_match
        item.setHidden(not should_show)
        
        # 자식이 매치되어서 내가 보여지는 경우, 나를 펼쳐야 자식이 보임
        if child_match:
            item.setExpanded(True)
            
        return should_show

    def show_product_list_context_menu(self, position):
        item = self.product_list.itemAt(position)
        if not item:
            return
            
        data = item.data(Qt.UserRole)
        # data is {'code': code, 'name': name, 'type': item_type}
        
        menu = QtWidgets.QMenu(self)
        where_used_action = menu.addAction("이 부품이 사용된 곳 보기 (Where Used)")
        where_used_action.triggered.connect(lambda: self.show_where_used(data))
        
        # ✅ [추가] 제품 마스터 바로가기
        menu.addSeparator()
        action_goto_master = menu.addAction("제품 마스터에서 보기 (수정)")
        action_goto_master.triggered.connect(lambda: self.go_to_product_master(data['code']))

        menu.exec_(self.product_list.mapToGlobal(position))

    def show_where_used(self, data):
        if not data: return
        # ✅ [수정] 모달리스로 변경 (.show() 사용)
        if hasattr(self, 'current_where_used_dialog') and self.current_where_used_dialog.isVisible():
            self.current_where_used_dialog.close()
            
        self.current_where_used_dialog = WhereUsedDialog(self, child_code=data['code'], child_name=data['name'])
        self.current_where_used_dialog.show()

    def set_current_parent(self, item):
        self.highlight_logic_check(item)

    def navigate_to_usage(self, parent_code, parent_name, child_code):
        """WhereUsedDialog에서 호출: 해당 부모를 로드하고 특정 자식을 하이라이트"""
        # 1. 부모가 현재 로드된 것과 다르면 로드
        if self.current_parent_code != parent_code:
            self.current_parent_code = parent_code
            self.current_parent_name = parent_name
            self.lbl_current_parent.setText(f"탐색됨: {parent_code} - {parent_name}")
            self.load_bom_tree()
        
        # 2. 자식 하이라이트
        # highlight_item_in_tree는 내부적으로 find_items_in_tree를 사용
        success = self.highlight_item_in_tree(child_code)
        
        if not success:
             QtWidgets.QMessageBox.warning(self, "알림", f"이 부모 품목({parent_code})의 BOM 트리 내에서\n자재 '{child_code}'를 찾을 수 없습니다.")

    def highlight_item_in_tree(self, code):
        """외부(예: WhereUsedDialog)에서 호출하여 특정 코드를 가진 아이템 하이라이트"""
        found_items = self.find_items_in_tree(code)
        if found_items:
            self.bom_tree.clearSelection()
            for found in found_items:
                found.setSelected(True)
                self.bom_tree.scrollToItem(found)
                parent = found.parent()
                while parent:
                    parent.setExpanded(True)
                    parent = parent.parent()
            return True
        else:
            return False

    def highlight_logic_check(self, item):
        """왼쪽 목록 더블클릭 핸들러: 트리에 있으면 하이라이트, 없으면 편집 모드로 진입"""
        data = item.data(Qt.UserRole)
        code = data['code']
        name = data['name']
        
        # 현재 트리에서 찾기
        found_items = self.find_items_in_tree(code)
        
        if found_items:
            # 찾았으면 하이라이트
            self.bom_tree.clearSelection()
            for found in found_items:
                found.setSelected(True)
                self.bom_tree.scrollToItem(found)
                parent = found.parent()
                while parent:
                    parent.setExpanded(True)
                    parent = parent.parent()
                    
            self.lbl_current_parent.setText(f"탐색됨: {code} (현재 BOM에 존재)")
        else:
            # 없으면 편집 모드로 진입 (기존 로직)
            self.current_parent_code = code
            self.current_parent_name = name
            self.lbl_current_parent.setText(f"편집 중: {self.current_parent_code} - {self.current_parent_name}")
            self.load_bom_tree()

    def find_items_in_tree(self, code):
        """트리 전체 순회하며 해당 코드 가진 아이템들 리스트 반환"""
        hits = []
        iterator = QtWidgets.QTreeWidgetItemIterator(self.bom_tree)
        while iterator.value():
            item = iterator.value()
            if item.text(1) == code: # col 1: code
                hits.append(item)
            iterator += 1
        return hits

    def load_bom_tree(self):
        """현재 부모의 BOM 트리 로드"""
        self.bom_tree.clear()
        if not self.current_parent_code:
            return

        tree_data = get_bom_tree(self.current_parent_code, max_depth=5)
        self.build_tree_items(self.bom_tree, tree_data)
        if self.target_search_bar.text(): 
            self.filter_bom_tree(self.target_search_bar.text())

    def build_tree_items(self, parent_widget, children_data):
        for idx, node in enumerate(children_data, start=1):
            item = QtWidgets.QTreeWidgetItem(parent_widget)
            
            # 0: No. (행 번호)
            item.setText(0, str(idx))
            
            # 1: Code
            item.setText(1, node['code'])
            
            # 2: Name
            item.setText(2, node['name'])
            
            # 3: Qty
            item.setText(3, str(node['qty']))

            # 4: Unit
            item.setText(4, node.get('unit', ''))
            
            # 5: Type (한글화)
            itype = node.get('item_type', '')
            type_map = {
                'PRODUCT': '완제품', 'MODULE': '모듈', 'PART': '부품',
                'SELLABLE': '완제품(구)', 'SUB_COMPONENT': '부품(구)'
            }
            item.setText(5, type_map.get(itype, itype))
            
            # 6: Remarks (비고)
            item.setText(6, node.get('remarks', ''))
            
            # ✅ [추가] 타입별 색상 강조
            # DB에 저장된 값이 'MODULE', 'Module', '모듈' 등 섞여 있을 수 있으므로 모두 체크
            upper_type = str(itype).upper()
            
            if upper_type in ('PRODUCT', 'SELLABLE', '완제품'):
                # 완제품: 연한 파랑 배경 + 굵은 글씨
                for c in range(item.columnCount()):
                    item.setBackground(c, QtGui.QColor("#E1F5FE"))
                # 품목코드(1), 품목명(2) 굵게
                font = item.font(1)
                font.setBold(True)
                item.setFont(1, font)
                item.setFont(2, font)
                
            elif upper_type in ('MODULE', '모듈'):
                # 모듈: 연한 주황 배경
                for c in range(item.columnCount()):
                    item.setBackground(c, QtGui.QColor("#FFF3E0"))
                # 모듈도 식별을 위해 굵게
                font = item.font(1)
                font.setBold(True)
                item.setFont(1, font)
                item.setFont(2, font)
            
            # 재귀적으로 자식 추가
            if node.get('children'):
                self.build_tree_items(item, node['children'])

    def eventFilter(self, source, event):
        """Drag & Drop 처리"""
        if source is self.bom_tree.viewport() and event.type() == QtCore.QEvent.Drop:
            if not self.current_parent_code:
                QtWidgets.QMessageBox.warning(self, "경고", "먼저 편집할 모품목을 선택(더블클릭)하세요.")
                return True
                
            mime_data = event.mimeData()
            # QListWidget의 기본 mimeData는 application/x-qabstractitemmodeldatalist 형식이지만
            # 직접 아이템을 가져오는게 더 쉽습니다.
            
            selected_items = self.product_list.selectedItems()
            if not selected_items:
                return True
                
            item = selected_items[0]
            data = item.data(Qt.UserRole)
            child_code = data['code']
            
            # 순환 참조 및 자기 자신 추가 방지 (DB단에서도 체크하지만 UI 반응성을 위해)
            if child_code == self.current_parent_code:
                QtWidgets.QMessageBox.warning(self, "오류", "자기 자신을 자식으로 추가할 수 없습니다.")
                return True
                
            # DB 추가 (기본 수량 1)
            success = add_bom_item(self.current_parent_code, child_code, 1.0)
            if success:
                self.load_bom_tree() # 새로고침
            else:
                QtWidgets.QMessageBox.warning(self, "오류", "BOM 추가 실패 (순환 참조 또는 DB 오류)")
                
            return True
            
        return super().eventFilter(source, event)

    def edit_item(self, item, column):
        """BOM 상세 편집 (소요량, 비고, 단위)"""
        child_code = item.text(1) # col 1: code
        current_qty_str = item.text(3) # col 3: qty
        current_unit = item.text(4) # col 4: unit
        current_remarks = item.text(6) # col 6: remarks (index shifted)
        
        try:
            current_qty = float(current_qty_str)
        except:
            current_qty = 1.0

        # 통합 편집 다이얼로그
        dialog = QtWidgets.QDialog(self)
        dialog.setWindowTitle(f"BOM 항목 수정 - {child_code}")
        dialog.resize(500, 400) # ✅ 크기 키움
        
        layout = QtWidgets.QFormLayout(dialog)
        layout.setFieldGrowthPolicy(QtWidgets.QFormLayout.AllNonFixedFieldsGrow)
        
        edt_qty = QtWidgets.QDoubleSpinBox()
        edt_qty.setRange(0.0, 100000.0)
        edt_qty.setDecimals(4)
        edt_qty.setValue(current_qty)

        edt_unit = QtWidgets.QLineEdit()
        edt_unit.setPlaceholderText("단위 (예: EA, M, kg...)")
        edt_unit.setText(current_unit)
        
        # ✅ QLineEdit -> QPlainTextEdit 변경 (여러 줄 입력)
        edt_remarks = QtWidgets.QPlainTextEdit()
        edt_remarks.setPlainText(current_remarks)
        edt_remarks.setPlaceholderText("비고 입력 (URL 등 긴 내용 가능)...")
        edt_remarks.setMinimumHeight(150) # 최소 높이 확보
        
        layout.addRow("소요량:", edt_qty)
        layout.addRow("단위:", edt_unit)
        layout.addRow("비고:", edt_remarks)
        
        btns = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel, Qt.Horizontal, dialog)
        btns.accepted.connect(dialog.accept)
        btns.rejected.connect(dialog.reject)
        layout.addRow(btns)
        
        if dialog.exec() == QtWidgets.QDialog.Accepted:
            new_qty = edt_qty.value()
            new_unit = edt_unit.text().strip()
            new_remarks = edt_remarks.toPlainText().strip() # ✅ toPlainText() 사용
            
            if update_bom_item(self.current_parent_code, child_code, new_qty=new_qty, new_remarks=new_remarks, new_unit=new_unit):
                item.setText(3, str(new_qty))
                item.setText(4, new_unit)
                item.setText(6, new_remarks)
            else:
                QtWidgets.QMessageBox.warning(self, "오류", "업데이트 실패")

    def show_context_menu(self, position):
        item = self.bom_tree.itemAt(position)
        if not item:
            return
            
        menu = QtWidgets.QMenu()
        action_edit = menu.addAction("수정 (소요량/비고)") # ✅ 추가
        action_edit.triggered.connect(lambda: self.edit_item(item, 0)) # 컬럼 무관하게 호출
        
        # ✅ [추가] 비고 복사 기능
        # ✅ [추가] 비고 복사 기능
        action_copy_desc = menu.addAction("비고를 품목 설명으로 복사")
        action_copy_desc.triggered.connect(self.copy_remarks_to_description)

        menu.addSeparator()

        # ✅ [추가] 이 부품이 사용된 곳 보기 (Where Used)
        action_where_used = menu.addAction("이 부품이 사용된 곳 보기 (Where Used)")
        action_where_used.triggered.connect(lambda: self.show_where_used_target(item))
        
        # ✅ [추가] 제품 마스터 바로가기
        action_goto_master = menu.addAction("제품 마스터에서 보기 (수정)")
        if item.text(1):
            action_goto_master.triggered.connect(lambda: self.go_to_product_master(item.text(1)))
        else:
            action_goto_master.setEnabled(False)

        menu.addSeparator()
        
        action_delete = menu.addAction("이 자식 품목 제거")
        action_delete.triggered.connect(lambda: self.delete_bom_item(item))
        menu.exec_(self.bom_tree.mapToGlobal(position))

    def delete_bom_item(self, item):
        child_code = item.text(1) # col 1: code
        confirm = QtWidgets.QMessageBox.question(
            self, "확인", f"정말 '{child_code}' 항목을 BOM에서 제거하시겠습니까?",
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No
        )
        if confirm == QtWidgets.QMessageBox.Yes:
            if delete_bom_item(self.current_parent_code, child_code):
                # UI에서 제거
                if item.parent():
                    item.parent().removeChild(item)
                else:
                    # 최상위 항목이면 (여기선 BOM 트리 루트의 직계 자식이 최상위)
                    index = self.bom_tree.indexOfTopLevelItem(item)
                    self.bom_tree.takeTopLevelItem(index)
                self.load_bom_tree() # 안전하게 전체 리로드
            else:
                QtWidgets.QMessageBox.warning(self, "오류", "삭제 실패")

    def show_where_used_target(self, item):
        """BOM 트리에서 선택된 항목의 'Where Used' 다이얼로그 표시"""
        code = item.text(1)
        name = item.text(2)
        item_code = item.text(1)
        item_name = item.text(2)
        if not item_code: return
        
        if hasattr(self, 'current_where_used_dialog') and self.current_where_used_dialog.isVisible():
            self.current_where_used_dialog.close()
            
        self.current_where_used_dialog = WhereUsedDialog(self, child_code=item_code, child_name=item_name, settings=self.settings)
        self.current_where_used_dialog.show()

    def copy_remarks_to_description(self):
        """선택된 BOM 항목들의 비고를 각 품목 마스터의 '설명'으로 일괄 복사"""
        selected_items = self.bom_tree.selectedItems()
        if not selected_items:
            return

        # 유효한 비고가 있는 항목만 필터링
        targets = []
        for item in selected_items:
            remarks = item.text(6).strip() # col 6: remarks
            if remarks:
                child_code = item.text(1) # col 1: code
                targets.append((child_code, remarks))
        
        if not targets:
            QtWidgets.QMessageBox.information(self, "알림", "선택된 항목 중 복사할 비고 내용이 있는 항목이 없습니다.")
            return

        count = len(targets)
        confirm = QtWidgets.QMessageBox.question(
            self, "확인", 
            f"선택된 {count}개 품목의 설명을\n현재 BOM 비고 내용으로 덮어쓰시겠습니까?",
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No
        )
        
        if confirm != QtWidgets.QMessageBox.Yes:
            return

        conn = get_conn()
        cur = conn.cursor()
        
        success_count = 0
        try:
            for child_code, remarks in targets:
                # 해당 코드의 최신(또는 활성) 품목 찾기
                cur.execute("SELECT id FROM product_master WHERE item_code = ? ORDER BY id DESC LIMIT 1", (child_code,))
                row = cur.fetchone()
                if row:
                    p_id = row[0]
                    cur.execute("UPDATE product_master SET description = ?, updated_at = datetime('now','localtime') WHERE id = ?", (remarks, p_id))
                    success_count += 1
            
            conn.commit()
            QtWidgets.QMessageBox.information(self, "완료", f"{success_count}건의 품목 설명이 업데이트되었습니다.")
            
        except Exception as e:
            conn.rollback()
            QtWidgets.QMessageBox.critical(self, "오류", f"일괄 업데이트 중 오류 발생: {e}")
        finally:
            conn.close()

    def on_btn_add_child_clicked(self):
        """[자식 추가] 버튼 핸들러"""
        if not self.current_parent_code:
            QtWidgets.QMessageBox.warning(self, "경고", "먼저 편집할 모품목을 선택(더블클릭)하세요.")
            return

        selected_items = self.product_list.selectedItems()
        if not selected_items:
            QtWidgets.QMessageBox.warning(self, "알림", "왼쪽 목록에서 추가할 자식 품목을 선택해주세요.")
            return
            
        item = selected_items[0]
        data = item.data(Qt.UserRole)
        child_code = data['code']
        
        # 순환 참조 방지
        if child_code == self.current_parent_code:
            QtWidgets.QMessageBox.warning(self, "오류", "자기 자신을 자식으로 추가할 수 없습니다.")
            return

        if add_bom_item(self.current_parent_code, child_code, 1.0):
             self.load_bom_tree()
        else:
             QtWidgets.QMessageBox.warning(self, "오류", "BOM 추가 실패 (순환 참조 또는 DB 오류)")

    def on_btn_manual_add_clicked(self):
        """[직접 입력] 버튼 핸들러"""
        if not self.current_parent_code:
            QtWidgets.QMessageBox.warning(self, "경고", "먼저 편집할 모품목을 선택(더블클릭)하세요.")
            return

        # 간단한 다이얼로그로 코드 입력 받기 (추후 검색 다이얼로그로 확장 가능)
        text, ok = QtWidgets.QInputDialog.getText(self, "품목 직접 추가", "추가할 품목 코드:")
        if ok and text:
            child_code = text.strip()
            if not child_code: return
            
            # 마스터에 존재하는지 확인 (선택사항, 없으면 자동등록 경고)
            # 여기서는 편의상 바로 추가 시도
            
            # 순환 참조 방지
            if child_code == self.current_parent_code:
                QtWidgets.QMessageBox.warning(self, "오류", "자기 자신을 자식으로 추가할 수 없습니다.")
                return

            # 추가 (기본 수량 1.0)
            if add_bom_item(self.current_parent_code, child_code, 1.0):
                self.load_bom_tree()
                # 마스터에 없는 경우 경고를 띄우거나 자동 등록 로직 (선택사항)
                # 여기서는 일단 성공 메시지 생략, 트리가 갱신됨.
            else:
                 QtWidgets.QMessageBox.warning(self, "오류", "추가 실패. 코드를 확인하거나 순환 참조를 점검하세요.")

    def save_column_widths(self):
        if not self.settings: return
        widths = []
        for col in range(self.bom_tree.columnCount()):
            widths.append(self.bom_tree.columnWidth(col))
        self.settings.setValue("bom_tree/column_widths", widths)

    def restore_column_widths(self):
        if not self.settings: return
        widths = self.settings.value("bom_tree/column_widths")
        if widths:
            for col, width in enumerate(widths):
                if col < self.bom_tree.columnCount():
                    self.bom_tree.setColumnWidth(col, int(width))

    def on_btn_remove_child_clicked(self):
        """[선택 삭제] 버튼 핸들러"""
        item = self.bom_tree.currentItem()
        if not item:
            QtWidgets.QMessageBox.warning(self, "알림", "삭제할 BOM 항목을 선택해주세요.")
            return
        self.delete_bom_item(item)

    def export_to_excel(self):
        if not self.current_parent_code:
             QtWidgets.QMessageBox.warning(self, "알림", "내보낼 BOM 트리가 없습니다.")
             return

        # 간단한 Flat Export 구현
        path, _ = QtWidgets.QFileDialog.getSaveFileName(self, "엑셀 내보내기", f"BOM_{self.current_parent_code}.xlsx", "Excel Files (*.xlsx)")
        if not path:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        # ✅ [수정] Unit, Remarks 추가
        ws.append(["Level", "Parent", "Child", "Name", "Qty", "Unit", "Type", "Remarks"])

        # Tree 순회
        iterator = QtWidgets.QTreeWidgetItemIterator(self.bom_tree)
        while iterator.value():
            item = iterator.value()
            
            # Level 계산
            level = 0
            p = item.parent()
            while p:
                level += 1
                p = p.parent()
                
            # Parent Code 찾기 (없으면 Root Parent)
            parent_code = item.parent().text(1) if item.parent() else self.current_parent_code # col 1: code
            
            code = item.text(1) # col 1
            name = item.text(2) # col 2
            qty = item.text(3)  # col 3
            unit = item.text(4) # col 4 ✅
            itype = item.text(5)# col 5 (index shifted due to unit)
            remarks = item.text(6) # col 6
            
            ws.append([level + 1, parent_code, code, name, qty, unit, itype, remarks])
            iterator += 1

        try:
            wb.save(path)
            QtWidgets.QMessageBox.information(self, "완료", "엑셀 내보내기 성공")
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "오류", f"엑셀 저장 실패: {e}")

    def import_from_excel(self):
        """엑셀 파일에서 BOM 구조 가져오기 (Parent, Child, Qty, Remarks)"""
        path, _ = QtWidgets.QFileDialog.getOpenFileName(self, "엑셀 가져오기", "", "Excel Files (*.xlsx *.xls)")
        if not path:
            return

        try:
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            
            success_count = 0
            fail_count = 0
            errors = []
            
            # 1. 헤더 파싱 (1행)
            headers = []
            for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
                headers = [str(h).strip().lower() if h else '' for h in row]
                break
                
            # 2. 컬럼 인덱스 찾기
            idx_parent = -1
            idx_child = -1
            idx_qty = -1
            idx_name = -1
            idx_type = -1
            idx_remarks = -1 
            idx_unit = -1 # ✅ [추가]
            
            # 헤더 기반 매핑 시도
            for i, h in enumerate(headers):
                # Parent
                if h in ['parent', 'parentcode', '모품목', '모품목코드', '부모']:
                    idx_parent = i
                # Child
                elif h in ['child', 'childcode', '자품목', '자품목코드', '자식']:
                    idx_child = i
                # Qty
                elif h in ['qty', 'quantity', 'soyou', '수량', '소요량']:
                    idx_qty = i
                # Name
                elif h in ['name', 'productname', '품목명', '제품명', '이름']:
                    idx_name = i
                # Type
                elif h in ['type', 'itemtype', '유형', '품목유형']:
                    idx_type = i
                # Remarks
                elif h in ['remarks', 'remark', '비고', 'memo', '메모']:
                    idx_remarks = i
                # Unit
                elif h in ['unit', 'danwi', '단위', 'uom']:
                    idx_unit = i
            
            # 헤더를 못 찾았거나 매핑이 불확실한 경우 기본값 적용
            if idx_parent == -1 or idx_child == -1:
                # 만약 첫 컬럼이 'level'이면 Export 된 파일로 간주 -> (1, 2, 4, 3, 5, 6) 적용
                if headers and 'level' in headers[0]:
                    idx_parent = 1
                    idx_child = 2
                    idx_qty = 4
                    idx_name = 3
                    idx_type = 5
                    idx_remarks = 6 
                    # unit은 export 포맷에 아직 없으므로 -1 유지 또는 필요시 추가
                else:
                    # 기본 포맷 [Parent, Child, Qty]
                    idx_parent = 0
                    idx_child = 1
                    idx_qty = 2
            
            # 2행부터 데이터 순회
            for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row:
                    continue
                    
                # 인덱스 범위 체크
                p_val = row[idx_parent] if idx_parent < len(row) else None
                c_val = row[idx_child] if idx_child < len(row) else None
                q_val = row[idx_qty] if (idx_qty != -1 and idx_qty < len(row)) else 1.0
                rem_val = row[idx_remarks] if (idx_remarks != -1 and idx_remarks < len(row)) else ""
                unit_val = row[idx_unit] if (idx_unit != -1 and idx_unit < len(row)) else ""
                
                parent_code = str(p_val).strip() if p_val else None
                child_code = str(c_val).strip() if c_val else None
                remarks = str(rem_val).strip() if rem_val else None
                unit = str(unit_val).strip() if unit_val else None
                
                try:
                    # ✅ 0도 유효한 값이므로 None이나 빈 문자열만 체크해야 함
                    if q_val is not None and str(q_val).strip() != "":
                         qty = float(q_val)
                    else:
                         qty = 1.0
                except ValueError:
                    qty = 1.0
                    
                if not parent_code or not child_code:
                    errors.append(f"Row {r_idx}: 필수 코드 누락 (Parent/Child 확인 필요)")
                    fail_count += 1
                    continue
                
                # [추가] 품목 마스터에 등록 (없으면 추가, 있으면 이름/타입 업데이트)
                try:
                    # Name 찾기
                    name = None
                    if idx_name != -1 and idx_name < len(row):
                        name = str(row[idx_name]).strip()
                        
                    # Type 찾기
                    itype = 'PART' # 기본값
                    if idx_type != -1 and idx_type < len(row):
                         itype = str(row[idx_type]).strip()
                         
                    # Child 품목 등록
                    if name:
                        # ✅ [수정] 판매가능 해제(0) 및 설명 문구 제거
                        add_or_update_product_master(
                            child_code, None, name, 0, 0, 
                            description="", 
                            item_type=itype,
                            is_active=0
                        )
                        
                except Exception as e:
                    # 마스터 등록 실패해도 BOM 추가는 시도 (이미 있을 수 있음)
                    print(f"Master Auto-Create Failed: {e}")

                # DB 추가
                try:
                    # ✅ remarks, unit 인자 추가
                    if add_bom_item(parent_code, child_code, qty, remarks, unit):
                        success_count += 1
                    else:
                        errors.append(f"Row {r_idx}: {parent_code}->{child_code} 추가 실패 (순환 참조 등)")
                        fail_count += 1
                except ValueError as ve:
                    errors.append(f"Row {r_idx}: {str(ve)}")
                    fail_count += 1
                except Exception as e:
                    errors.append(f"Row {r_idx}: 알 수 없는 오류 - {str(e)}")
                    fail_count += 1
            
            msg = f"가져오기 완료\n성공: {success_count}건\n실패/스킵: {fail_count}건"
            if errors:
                msg += "\n\n[오류 상세 (상위 10건)]\n" + "\n".join(errors[:10])
                if len(errors) > 10:
                    msg += "\n... (더 많은 오류가 있음)"
            
            if fail_count > 0:
                 QtWidgets.QMessageBox.warning(self, "가져오기 결과", msg)
            else:
                 QtWidgets.QMessageBox.information(self, "완료", msg)
            
            # 현재 뷰 갱신
            if self.current_parent_code:
                self.load_bom_tree()
                
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "오류", f"엑셀 가져오기 실패: {e}")

    def go_to_product_master(self, item_code):
        """특정 품목 코드를 가지고 제품 마스터 탭으로 이동"""
        if not item_code:
            return

        # 1. Main Window 찾기
        main_window = self.window()
        
        if not hasattr(main_window, 'main_tabs'):
            print("Main Window main_tabs not found")
            return

        tabs = main_window.main_tabs
        
        # 2. 제품 마스터 탭 찾기
        target_index = -1
        product_master_widget = None
        
        for i in range(tabs.count()):
            widget = tabs.widget(i)
            # ProductMasterWidget 인스턴스인지 확인
            if isinstance(widget, ProductMasterWidget):
                target_index = i
                product_master_widget = widget
                break
        
        if target_index != -1 and product_master_widget:
            # 3. 대상 제품 정보 조회 (상태, 유형 확인)
            try:
                conn = get_conn()
                cur = conn.cursor()
                cur.execute("SELECT is_active, item_type FROM product_master WHERE item_code = ?", (item_code,))
                row = cur.fetchone()
                conn.close()
                
                is_active = 1
                item_type = 'PART'
                if row:
                    is_active = row[0]
                    item_type = row[1]
            except Exception as e:
                print(f"Product Lookup Error: {e}")
                is_active = 1
                item_type = 'PART'

            # 4. ProductMasterWidget 필터 강제 조정 (시그널 블록)
            # - 단종품이면 '전체 보기' 켜기
            if not is_active and not product_master_widget.show_all:
                product_master_widget.show_all = True
                product_master_widget.btn_show_all.setChecked(True)
                product_master_widget.btn_show_all.setText("전체보기")
            
            # - 해당 유형이 꺼져 있으면 켜기
            if item_type and item_type not in product_master_widget.selected_types:
                product_master_widget.selected_types.add(item_type)
                product_master_widget.update_filter_button_text()
                # 메뉴 액션 상태 동기화
                if item_type in product_master_widget.filter_actions:
                    product_master_widget.filter_actions[item_type].setChecked(True)

            # 5. 검색창에 코드 입력 (가장 확실한 필터링 방법)
            # 시그널을 블록하여 중복 로드를 방지하고 마지막에 한 번 로드
            product_master_widget.search_bar.blockSignals(True)
            product_master_widget.search_bar.setText(item_code)
            product_master_widget.search_bar.blockSignals(False)

            # 6. 하이라이트 타겟 설정
            product_master_widget.current_product_code = item_code
            
            # 7. 탭 전환
            tabs.setCurrentIndex(target_index)
            
            # 8. 데이터 로드 (필터 변경사항 반영 및 하이라이트 수행)
            product_master_widget.load_product_list()